import streamlit as st
import pandas as pd
import json
import os
from datetime import datetime, timedelta
import openpyxl
import re

# --- 設定・初期化 ---
st.set_page_config(page_title="禅道会Tシャツ在庫管理", layout="wide")

DATA_FILE = "inventory_db.json"
TSHIRT_TYPES = [
    'パンクラス×禅道会コラボTシャツ(ホワイト)ゼンプロマークなし',
    'パンクラス×禅道会コラボTシャツ(ブラック)ゼンプロマークなし',
    'パンクラス×禅道会コラボTシャツ(ホワイト)ゼンプロマークあり',
    'パンクラス×禅道会コラボTシャツ(ブラック)ゼンプロマークあり'
]
SIZES = ['150cm', '160cm', 'S', 'M', 'L', 'XL', 'XXL']

# --- データ操作関数 ---
def load_all_records():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_all_records(records):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def normalize_size(val):
    val = str(val).strip().upper()
    if '150' in val: return '150cm'
    if '160' in val: return '160cm'
    if 'XXL' in val or '3L' in val: return 'XXL'
    if 'XL' in val or 'LL' in val: return 'XL'
    if 'L' in val: return 'L'
    if 'M' in val: return 'M'
    if 'S' in val: return 'S'
    return None

def determine_type(filename):
    is_white = '白' in filename or 'ホワイト' in filename
    is_ari = 'あり' in filename
    if is_white and not is_ari: return TSHIRT_TYPES[0]
    if not is_white and not is_ari: return TSHIRT_TYPES[1]
    if is_white and is_ari: return TSHIRT_TYPES[2]
    if not is_white and is_ari: return TSHIRT_TYPES[3]
    return None

# --- メインUI ---
st.title("📦 パンクラス×禅道会 在庫管理システム")

# セッション状態の初期化
if 'records' not in st.session_state:
    st.session_state.records = load_all_records()

# サイドバー：Excelインポート
with st.sidebar:
    st.header("📥 Excel取込")
    uploaded_files = st.file_uploader("在庫管理表を選択(複数可)", type=['xlsx'], accept_multiple_files=True)
    
    if st.button("Excelデータを反映する") and uploaded_files:
        temp_records = {r['date']: r for r in st.session_state.records}
        
        for uploaded_file in uploaded_files:
            ttype = determine_type(uploaded_file.name)
            if not ttype: continue
            
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            
            # 日付行とデータ行の解析
            header_row = None
            date_cols = {}
            for r in range(1, 10):
                row_vals = [c.value for c in ws[r]]
                for idx, v in enumerate(row_vals):
                    if isinstance(v, datetime):
                        date_cols[idx] = v.strftime('%Y-%m-%d')
                        header_row = r
                if date_cols: break
            
            if header_row:
                for r in range(header_row + 1, ws.max_row + 1):
                    p_name = ws.cell(row=r, column=2).value # 通常B列
                    size = normalize_size(p_name)
                    if not size: continue
                    
                    for col_idx, d_str in date_cols.items():
                        count = ws.cell(row=r, column=col_idx + 1).value or 0
                        if d_str not in temp_records:
                            temp_records[d_str] = {
                                "date": d_str,
                                "inventory": {t: {s: 0 for s in SIZES} for t in TSHIRT_TYPES}
                            }
                        temp_records[d_str]["inventory"][ttype][size] = int(count)
        
        st.session_state.records = sorted(list(temp_records.values()), key=lambda x: x['date'], reverse=True)
        save_all_records(st.session_state.records)
        st.success("インポート完了！")

# タブ分け
tab1, tab2 = st.tabs(["📝 今日の在庫入力", "📊 履歴・グラフ"])

with tab1:
    st.subheader("本日の在庫数を確認・修正してください")
    target_date = st.date_input("記録日", datetime.now()).strftime('%Y-%m-%d')
    
    # 既存データの検索
    current_data = next((r for r in st.session_state.records if r['date'] == target_date), None)
    
    # データがない場合は前日のデータをコピー
    if not current_data and st.session_state.records:
        prev_data = st.session_state.records[0] # 最新のもの
        current_data = {"date": target_date, "inventory": json.loads(json.dumps(prev_data['inventory']))}
    elif not current_data:
        current_data = {"date": target_date, "inventory": {t: {s: 0 for s in SIZES} for t in TSHIRT_TYPES}}

    # 入力フォーム
    new_inventory = {}
    cols = st.columns(2)
    for idx, ttype in enumerate(TSHIRT_TYPES):
        with cols[idx % 2]:
            st.markdown(f"**{ttype}**")
            new_inventory[ttype] = {}
            # 1行にサイズを並べる
            size_cols = st.columns(len(SIZES))
            for s_idx, size in enumerate(SIZES):
                with size_cols[s_idx]:
                    val = current_data['inventory'].get(ttype, {}).get(size, 0)
                    new_inventory[ttype][size] = st.number_input(f"{size}", min_value=0, value=val, key=f"{target_date}{ttype}{size}")

    if st.button("この内容で保存する", type="primary"):
        # 既存リストを更新
        updated_records = [r for r in st.session_state.records if r['date'] != target_date]
        updated_records.append({"date": target_date, "inventory": new_inventory})
        st.session_state.records = sorted(updated_records, key=lambda x: x['date'], reverse=True)
        save_all_records(st.session_state.records)
        st.balloons()
        st.success(f"{target_date} のデータを保存しました。")

with tab2:
    st.subheader("在庫推移・履歴")
    if not st.session_state.records:
        st.info("データがまだありません。")
    else:
        # 表形式で表示
        history_df = []
        for r in st.session_state.records:
            for ttype in TSHIRT_TYPES:
                row = {"日付": r['date'], "種類": ttype}
                row.update(r['inventory'][ttype])
                row["合計"] = sum(r['inventory'][ttype].values())
                history_df.append(row)
        
        df = pd.DataFrame(history_df)
        
        selected_type = st.selectbox("種類で絞り込み", ["すべて"] + TSHIRT_TYPES)
        display_df = df if selected_type == "すべて" else df[df["種類"] == selected_type]
        
        st.dataframe(display_df, use_container_width=True)
        
        # 簡易グラフ
        if selected_type != "すべて":
            st.line_chart(display_df.set_index("日付")[SIZES])